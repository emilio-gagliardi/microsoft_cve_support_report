import os
import pandas as pd
from typing import Any, Dict
from kedro.io import AbstractDataSet
from pathlib import Path
import openpyxl
import re
# reading xlsx files requires openpyxl


class LatestExcelDataSet(AbstractDataSet):
    def __init__(self, filepath: str, load_args: Dict[str, Any] = None):
        """Initialize the dataset to load the latest Excel file based on its naming pattern."""
        super().__init__()
        self.filepath = Path(filepath)
        self.load_args = load_args if load_args is not None else {}

    def _load(self) -> pd.DataFrame:
        files = list(self.filepath.glob('*.xlsx'))
        if not files:
            raise FileNotFoundError(f"No Excel files found in directory {self.filepath}")
        latest_file = max(files, key=os.path.getctime)

        wb = openpyxl.load_workbook(latest_file)
        sheet = wb.active

        data = []
        urls = [[] for _ in range(sheet.max_row - 1)]
        hyperlink_columns = set()

        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row)):
            row_data = []
            for j, cell in enumerate(row):
                row_data.append(cell.value)
                if cell.hyperlink:
                    urls[i].append(cell.hyperlink.target)
                    hyperlink_columns.add(j)
                elif j in hyperlink_columns:
                    urls[i].append(None)
            data.append(row_data)

        column_names = [cell.value for cell in sheet[1]]
        df = pd.DataFrame(data, columns=column_names)

        if hyperlink_columns:
            df_urls = pd.DataFrame(urls, columns=[f'{column_names[j]}_URL' for j in range(len(column_names)) if j in hyperlink_columns])
            df = pd.concat([df, df_urls], axis=1)
        else:
            print("Custom Dataset found no links in excel file")

        # Convert 'Release date' string to datetime
        release_date_column_name = 'Release date'
        df[release_date_column_name] = df[release_date_column_name].apply(self.parse_dates)
        # print("Sample data after conversion:", df[release_date_column_name].head())

        # New logic to check 'kb_id' format and create custom 'kb_id' if necessary
        kb_id_column_name = 'Article'  # The column name for kb_id
        cve_id_column_name = 'Details'  
        build_number_column_name = 'Build Number'
        # Check if 'kb_id', 'cve_id', 'build_number', and 'Release date' columns exist
        # Check if all required columns exist in the DataFrame
        if (kb_id_column_name in df.columns and cve_id_column_name in df.columns and build_number_column_name in df.columns and release_date_column_name in df.columns):

            # Apply a lambda function to each row in the DataFrame
            df[kb_id_column_name] = df.apply(
                lambda row: (
                    # Create a custom 'kb_id' by concatenating 'build_number', 'cve_id', and 'release_date'
                    f"{row[build_number_column_name]}_{row[cve_id_column_name]}_{row[release_date_column_name].strftime('%Y-%m-%d')}"
                    # This concatenation happens only if:
                    # 1. 'kb_id' does not already match a 7-digit pattern
                    # 2. 'cve_id' is not empty or NaN
                    # 3. 'release_date' is not NaT (not a date)
                    if (not re.match(r'^\d{7,}$', str(row[kb_id_column_name])) and 
                        row[cve_id_column_name] and 
                        pd.notnull(row[release_date_column_name]))
                    # Otherwise, keep the existing 'kb_id'
                    else row[kb_id_column_name]
                ),
                axis=1  # Apply the lambda function row-wise
            )

        else:
            print(f"Something didn't work with these columns: {kb_id_column_name}, {cve_id_column_name}, {build_number_column_name}, {release_date_column_name}. Compared to df.columns: {df.columns}")

        return df


    def _save(self, data: pd.DataFrame) -> None:
        """Save method is not implemented as this dataset is read-only."""
        raise NotImplementedError("Save method of LatestExcelDataSet is not implemented.")


    def _describe(self) -> Dict[str, Any]:
        """Return a dict that describes the attributes of the dataset."""
        return dict(filepath=self.filepath, load_args=self.load_args)
    

    def parse_dates(self, date_str):
        for fmt in ('%d %b %Y', '%b %d, %Y'):
            try:
                return pd.to_datetime(date_str, format=fmt)
            except ValueError:
                continue
        return pd.NaT
